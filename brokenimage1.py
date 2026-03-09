# # fast_broken_images_crawler.py

# import time
# import requests
# import openpyxl
# from urllib.parse import urljoin, urlparse
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from webdriver_manager.chrome import ChromeDriverManager
# from collections import deque
# from concurrent.futures import ThreadPoolExecutor


# # Device viewports
# DEVICES = {
#     "Desktop": (1920, 1080),
#      "laptops":(1025,1420),
#     "iPad": (768, 1024),
#     "Mobile": (375, 667),
# }

# START_URL = "https://technologyrental.sg/"   # change to your website
# MAX_PAGES = 300                     # audit limit
# MAX_WORKERS = 20                    # parallel HTTP requests
# TIMEOUT = 3                         # HTTP timeout in seconds


# def setup_driver():
#     """Launch Chrome once and reuse by resizing."""
#     chrome_options = Options()
#     chrome_options.add_argument("--disable-infobars")
#     chrome_options.add_argument("--disable-extensions")
#     chrome_options.add_argument("--disable-gpu")
#     driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
#     return driver


# def check_image_status(src):
#     """Check if image URL is broken."""
#     if not src:
#         return "Missing src"
#     try:
#         response = requests.get(src, timeout=TIMEOUT)
#         if response.status_code == 200:
#             return "OK"
#         else:
#             return f"Broken (Status {response.status_code})"
#     except Exception as e:
#         return f"Error: {str(e)}"


# def check_images(driver, device):
#     """Check all images on current page using threads."""
#     results = []
#     images = driver.find_elements(By.TAG_NAME, "img")
#     img_srcs = [img.get_attribute("src") for img in images]

#     with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
#         statuses = list(executor.map(check_image_status, img_srcs))

#     for src, status in zip(img_srcs, statuses):
#         results.append((device, driver.current_url, src if src else "N/A", status))

#     return results


# def get_links(driver, base_url):
#     """Extract internal links from current page."""
#     links = set()
#     for el in driver.find_elements(By.TAG_NAME, "a"):
#         href = el.get_attribute("href")
#         if href and href.startswith(base_url):
#             links.add(href.split("#")[0])  # remove anchors
#     return links


# def save_to_excel(results, filename="broken_images_fast_audit.xlsx"):
#     """Save results to Excel file."""
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Broken Images Report"

#     headers = ["Device", "Page URL", "Image URL", "Status"]
#     ws.append(headers)

#     for row in results:
#         ws.append(row)

#     wb.save(filename)
#     print(f"✅ Report saved as {filename}")


# def crawl_site():
#     all_results = []
#     visited = set()
#     queue = deque([START_URL])

#     base_url = urlparse(START_URL).scheme + "://" + urlparse(START_URL).netloc
#     driver = setup_driver()

#     while queue and len(visited) < MAX_PAGES:
#         url = queue.popleft()
#         if url in visited:
#             continue
#         visited.add(url)

#         print(f"\n🌍 Crawling {url} ({len(visited)}/{MAX_PAGES})")
#         try:
#             driver.get(url)
#             time.sleep(2)  # short wait

#             for device, (width, height) in DEVICES.items():
#                 driver.set_window_size(width, height)
#                 time.sleep(0.5)  # allow resize
#                 results = check_images(driver, device)
#                 all_results.extend(results)

#             # collect links (desktop only to avoid duplicates)
#             links = get_links(driver, base_url)
#             for link in links:
#                 if link not in visited and len(visited) + len(queue) < MAX_PAGES:
#                     queue.append(link)

#         except Exception as e:
#             print(f"⚠️ Error on {url}: {e}")

#     driver.quit()
#     save_to_excel(all_results)


# if __name__ == "__main__":
#     crawl_site()















# import time
# import requests
# import openpyxl
# from urllib.parse import urlparse
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from webdriver_manager.chrome import ChromeDriverManager
# from collections import deque
# from concurrent.futures import ThreadPoolExecutor

# # --------------------------
# # CONFIGURATION
# # --------------------------
# DEVICES = {
#     "Desktop": (1920, 1080),
#     "Laptop": (1366, 768),
#     "iPad": (768, 1024),
#     "Mobile": (375, 667),
# }

# START_URL = "https://technologyrental.sg/"  # Change to your website
# MAX_PAGES = 200                             # Max pages to crawl
# MAX_WORKERS = 40                            # More threads = faster image checking
# TIMEOUT = 3                                 # HTTP request timeout in seconds
# PAGE_WAIT = 1.5                             # Time to wait for page load

# # --------------------------
# # DRIVER SETUP
# # --------------------------
# def setup_driver():
#     """Initialize Chrome browser (visible and faster)."""
#     chrome_options = Options()
#     chrome_options.add_argument("--disable-infobars")
#     chrome_options.add_argument("--disable-extensions")
#     chrome_options.add_argument("--start-maximized")
#     chrome_options.add_argument("--no-sandbox")
#     chrome_options.add_argument("--disable-dev-shm-usage")
#     # REMOVE headless so browser opens visibly
#     driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
#     return driver

# # --------------------------
# # IMAGE CHECKING
# # --------------------------
# def check_image_status(src):
#     """Check if image URL is valid or broken."""
#     if not src:
#         return "Missing src"
#     try:
#         response = requests.head(src, timeout=TIMEOUT)
#         if response.status_code == 405:
#             response = requests.get(src, timeout=TIMEOUT)
#         if response.status_code == 200:
#             return "OK"
#         return f"Broken (Status {response.status_code})"
#     except Exception as e:
#         return f"Error: {str(e)}"

# def extract_image_sources(driver):
#     """
#     Extract all image URLs including lazy-loaded ones.
#     Looks for: src, data-src, data-lazy, data-original, srcset, data-srcset.
#     """
#     images = driver.find_elements(By.TAG_NAME, "img")
#     img_urls = set()

#     for img in images:
#         possible_attrs = [
#             img.get_attribute("src"),
#             img.get_attribute("data-src"),
#             img.get_attribute("data-lazy"),
#             img.get_attribute("data-original"),
#             img.get_attribute("data-srcset"),
#             img.get_attribute("srcset"),
#         ]

#         for attr in possible_attrs:
#             if attr and attr.strip():
#                 attr = attr.strip()
#                 if attr.startswith("//"):
#                     attr = "https:" + attr
#                 elif attr.startswith("/"):
#                     attr = driver.current_url.rstrip("/") + attr
#                 if attr.startswith("http"):
#                     img_urls.add(attr)
#                 break  # Use first valid URL

#     return list(img_urls)

# def check_images(driver, device):
#     """Check all images on a page for a specific device viewport."""
#     results = []
#     img_srcs = extract_image_sources(driver)

#     if not img_srcs:
#         return [(device, driver.current_url, "No Images Found", "N/A")]

#     with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
#         statuses = list(executor.map(check_image_status, img_srcs))

#     for src, status in zip(img_srcs, statuses):
#         results.append((device, driver.current_url, src, status))
#     return results

# # --------------------------
# # LINK COLLECTION
# # --------------------------
# def get_links(driver, base_url):
#     """Extract all internal links on a page."""
#     links = set()
#     for el in driver.find_elements(By.TAG_NAME, "a"):
#         href = el.get_attribute("href")
#         if href and href.startswith(base_url):
#             links.add(href.split("#")[0])
#     return links

# # --------------------------
# # SAVE TO EXCEL
# # --------------------------
# def save_to_excel(results, filename="broken_images_fast_audit.xlsx"):
#     """Save results to an Excel report."""
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Broken Images Report"

#     headers = ["Device", "Page URL", "Image URL", "Status"]
#     ws.append(headers)
#     for row in results:
#         ws.append(row)

#     wb.save(filename)
#     print(f"✅ Report saved as {filename}")

# # --------------------------
# # MAIN CRAWLER
# # --------------------------
# def crawl_site():
#     """Main function to crawl pages and check images."""
#     all_results = []
#     visited = set()
#     queue = deque([START_URL])
#     base_url = urlparse(START_URL).scheme + "://" + urlparse(START_URL).netloc
#     driver = setup_driver()

#     print("🚀 Starting Broken Image Crawler...\n")

#     while queue and len(visited) < MAX_PAGES:
#         url = queue.popleft()
#         if url in visited:
#             continue
#         visited.add(url)

#         print(f"🌍 Crawling {url} ({len(visited)}/{MAX_PAGES})")

#         try:
#             driver.get(url)
#             time.sleep(PAGE_WAIT)

#             for device, (width, height) in DEVICES.items():
#                 driver.set_window_size(width, height)
#                 time.sleep(0.5)
#                 results = check_images(driver, device)
#                 all_results.extend(results)

#             links = get_links(driver, base_url)
#             for link in links:
#                 if link not in visited and len(visited) + len(queue) < MAX_PAGES:
#                     queue.append(link)

#         except Exception as e:
#             print(f"⚠️ Error on {url}: {e}")

#     driver.quit()
#     save_to_excel(all_results)
#     print("\n🎯 Crawl completed successfully!")

# # --------------------------
# # RUN
# # --------------------------
# if __name__ == "__main__":
#     crawl_site()







import time
import requests
import openpyxl
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from requests.adapters import HTTPAdapter, Retry

# --------------------------
# CONFIGURATION
# --------------------------
DEVICES = {
    "Desktop": (1920, 1080),
    "Laptop": (1366, 768),
    "iPad": (768, 1024),
    "Mobile": (375, 667),
}

START_URL = "https://oneworldrental.ae/"  # Change to your website
MAX_PAGES = 200                              # Maximum pages to crawl
MAX_WORKERS = 40                             # Threads for parallel image checks
TIMEOUT = 10                                 # Increased timeout for stability
PAGE_WAIT = 2                                # Time to wait for page load

# --------------------------
# GLOBAL SESSION WITH RETRIES
# --------------------------
session = requests.Session()
retries = Retry(
    total=3,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
)
adapter = HTTPAdapter(max_retries=retries)
session.mount("http://", adapter)
session.mount("https://", adapter)

# --------------------------
# DRIVER SETUP
# --------------------------
def setup_driver():
    """Initialize Chrome browser (visible and optimized)."""
    chrome_options = Options()
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # Visible browser for debugging
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    return driver

# --------------------------
# IMAGE CHECKING
# --------------------------
def check_image_status(src):
    """Check if an image URL is valid, with retry and timeout."""
    if not src:
        return "Missing src"

    try:
        response = session.head(src, timeout=TIMEOUT)
        if response.status_code == 405:  # HEAD not allowed
            response = session.get(src, timeout=TIMEOUT)
        if response.status_code == 200:
            return "OK"
        return f"Broken (Status {response.status_code})"
    except requests.exceptions.RequestException as e:
        return f"Error: {str(e)}"

# --------------------------
# IMAGE EXTRACTION (with lazy-load handling)
# --------------------------
def extract_image_sources(driver):
    """
    Extract all image URLs including lazy-loaded ones.
    Supports: src, data-src, data-lazy, data-original, srcset, data-srcset.
    """
    # Scroll page to load lazy images
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    images = driver.find_elements(By.TAG_NAME, "img")
    img_urls = set()

    for img in images:
        possible_attrs = [
            img.get_attribute("src"),
            img.get_attribute("data-src"),
            img.get_attribute("data-lazy"),
            img.get_attribute("data-original"),
            img.get_attribute("data-srcset"),
            img.get_attribute("srcset"),
        ]

        for attr in possible_attrs:
            if attr and attr.strip():
                attr = attr.strip()
                if attr.startswith("//"):
                    attr = "https:" + attr
                elif attr.startswith("/"):
                    attr = driver.current_url.rstrip("/") + attr
                if attr.startswith("http"):
                    img_urls.add(attr)
                break  # Use first valid URL

    return list(img_urls)

def check_images(driver, device):
    """Check all images on the page for the given device."""
    results = []
    img_srcs = extract_image_sources(driver)

    if not img_srcs:
        return [(device, driver.current_url, "No Images Found", "N/A")]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        statuses = list(executor.map(check_image_status, img_srcs))

    for src, status in zip(img_srcs, statuses):
        results.append((device, driver.current_url, src, status))
    return results

# --------------------------
# LINK COLLECTION
# --------------------------
def get_links(driver, base_url):
    """Extract all internal links from the current page."""
    links = set()
    for el in driver.find_elements(By.TAG_NAME, "a"):
        href = el.get_attribute("href")
        if href and href.startswith(base_url):
            links.add(href.split("#")[0])
    return links

# --------------------------
# SAVE TO EXCEL
# --------------------------
def save_to_excel(results, filename="broken_images_fast_audit.xlsx"):
    """Save the results to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Broken Images Report"
    headers = ["Device", "Page URL", "Image URL", "Status"]
    ws.append(headers)
    for row in results:
        ws.append(row)
    wb.save(filename)
    print(f"✅ Report saved as {filename}")

# --------------------------
# MAIN CRAWLER
# --------------------------
def crawl_site():
    """Main crawler to audit website images across devices."""
    all_results = []
    visited = set()
    queue = deque([START_URL])
    base_url = urlparse(START_URL).scheme + "://" + urlparse(START_URL).netloc
    driver = setup_driver()

    print("🚀 Starting Broken Image Crawler...\n")

    while queue and len(visited) < MAX_PAGES:
        url = queue.popleft()
        if url in visited:
            continue
        visited.add(url)

        print(f"🌍 Crawling {url} ({len(visited)}/{MAX_PAGES})")

        try:
            driver.get(url)
            time.sleep(PAGE_WAIT)

            for device, (width, height) in DEVICES.items():
                driver.set_window_size(width, height)
                time.sleep(0.5)
                results = check_images(driver, device)
                all_results.extend(results)

            links = get_links(driver, base_url)
            for link in links:
                if link not in visited and len(visited) + len(queue) < MAX_PAGES:
                    queue.append(link)

        except Exception as e:
            print(f"⚠️ Error on {url}: {e}")

    driver.quit()
    save_to_excel(all_results)
    print("\n🎯 Crawl completed successfully!")

# --------------------------
# RUN SCRIPT
# --------------------------
if __name__ == "__main__":
    crawl_site()
