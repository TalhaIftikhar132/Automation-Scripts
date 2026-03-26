
"""
TEAMS Europe Participant Scraper (Selenium)
============================================
Extracts ONLY the specified matchmaking fields from each profile.
Each field becomes its own separate column in the Excel output.

Requirements:
    pip install selenium openpyxl webdriver-manager

Usage:
    python teamseurope_selenium_scraper.py

Output:
    teamseurope_participants.xlsx
"""

import time
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── URLs ──────────────────────────────────────────────────────────────────────
URLS = [
    "https:Example of site page",




]

OUTPUT_FILE = "teamseurope_participants.xlsx"
LOGIN_WAIT  = 30  # seconds you have to log in manually if needed
PAGE_WAIT   = 8   # max seconds to wait for each page to load

# ── Exact field labels as they appear on the page ────────────────────────────
# Each becomes its own column. Order is preserved in Excel.
FIELDS = [
    "First Name",
    "Last Name",
    "Company",
    "Job Title",
    "Which category best represents your organisation?",
    "How many events do you typically plan a year?",
    "What types of event do you organise?",
    "Which locations are you interested in for your events?",
    "Regions of Interest UK",
    "Countries of Interest - Europe",
    "What is your average room rate per person per night?",
    "Please estimate how many total attendees participate in the events you plan annually?",
    "If looking for a hotel/meeting facility, what is the largest number of breakout rooms that you need",
    "If looking for a hotel/meeting facility, what is the largest venue space you need for each event",
    "For your open events, which type of accommodations do you use? Check all that apply",
    "Which other venues/services do you typically use? Check all that apply",
]

# Shorter display headers for the Excel columns (same order as FIELDS)
DISPLAY_HEADERS = [
    "First Name",
    "Last Name",
    "Company",
    "Job Title",
    "Organisation Category",
    "Events Per Year",
    "Event Types",
    "Locations of Interest",
    "UK Regions of Interest",
    "European Countries of Interest",
    "Avg Room Rate (per person/night)",
    "Total Annual Attendees",
    "Max Breakout Rooms Needed",
    "Max Venue Space Needed",
    "Accommodation Types Used",
    "Venues / Services Used",
]

ALL_COLUMNS    = ["Profile ID"] + FIELDS
ALL_HEADERS    = ["Profile ID"] + DISPLAY_HEADERS


# ── Browser Setup ─────────────────────────────────────────────────────────────
def create_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


# ── Parser ────────────────────────────────────────────────────────────────────
def clean(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def parse_profile(driver: webdriver.Chrome, url: str) -> dict:
    pid = url.rstrip("/").split("/")[-1]

    # All fields default to empty string
    record = {col: "" for col in ALL_COLUMNS}
    record["Profile ID"] = pid

    try:
        driver.get(url)
        WebDriverWait(driver, PAGE_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(2.5)  # allow JS rendering to finish
    except Exception as e:
        print(f"    [WARN] Load error: {e}")
        return record

    # ── Strategy 1: <dt> / <dd> definition list pairs ────────────────────────
    try:
        dts = driver.find_elements(By.TAG_NAME, "dt")
        dds = driver.find_elements(By.TAG_NAME, "dd")
        for dt, dd in zip(dts, dds):
            label = clean(dt.text)
            value = clean(dd.text)
            for field in FIELDS:
                if label.lower() == field.lower() or label.lower().startswith(field.lower()[:45]):
                    if not record[field]:
                        record[field] = value
                    break
    except Exception:
        pass

    # ── Strategy 2: Two-cell table rows ───────────────────────────────────────
    try:
        rows = driver.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) == 2:
                label = clean(cells[0].text)
                value = clean(cells[1].text)
                for field in FIELDS:
                    if label.lower() == field.lower() or label.lower().startswith(field.lower()[:45]):
                        if not record[field]:
                            record[field] = value
                        break
    except Exception:
        pass

    # ── Strategy 3: Full page text, label then value on next line ─────────────
    # Used as a fallback for any field still empty after strategies 1 & 2
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        lines     = [clean(l) for l in body_text.splitlines() if clean(l)]

        for i, line in enumerate(lines):
            for field in FIELDS:
                if record[field]:
                    continue  # already filled by an earlier strategy
                # Match if the line equals the label OR starts with the first 45 chars of it
                if line.lower() == field.lower() or line.lower().startswith(field.lower()[:45]):
                    # Walk forward to find the value
                    for j in range(i + 1, min(i + 6, len(lines))):
                        candidate = lines[j]
                        # Skip if it looks like another field label
                        is_label = any(
                            candidate.lower().startswith(f.lower()[:30])
                            for f in FIELDS
                        )
                        if candidate and not is_label:
                            record[field] = candidate
                            break
                    break  # stop checking other fields for this line
    except Exception:
        pass

    return record


# ── Excel Writer ──────────────────────────────────────────────────────────────
def save_excel(records: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    # --- Styles ---
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", start_color="DCE6F1")
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin   = Side(style="thin", color="B8B8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header row ---
    for ci, header in enumerate(ALL_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    # --- Data rows ---
    for ri, rec in enumerate(records, 2):
        for ci, col in enumerate(ALL_COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=rec.get(col, ""))
            cell.font      = cell_font
            cell.alignment = cell_align
            cell.border    = border
            if ri % 2 == 0:
                cell.fill = alt_fill

    # --- Column widths (by position) ---
    widths = [12, 14, 14, 28, 30, 28, 18, 30, 28, 32, 35, 28, 22, 22, 26, 35, 38]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 50
    ws.freeze_panes = "B2"

    wb.save(path)
    print(f"\n✅  Saved {len(records)} records → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  TEAMS Europe Participant Scraper")
    print("=" * 60)
    print(f"\n  {len(URLS)} profiles  |  {len(FIELDS)} fields each")
    print(f"  Output file → {OUTPUT_FILE}\n")

    driver = create_driver()

    driver.get("https://app.teamseurope.com/newfront/")
    print(f"⏳  Browser opened. You have {LOGIN_WAIT}s to log in if prompted …")
    time.sleep(LOGIN_WAIT)

    records = []
    for i, url in enumerate(URLS, 1):
        pid = url.rstrip("/").split("/")[-1]
        print(f"  [{i:02d}/{len(URLS)}] ID {pid} … ", end="", flush=True)
        rec = parse_profile(driver, url)
        records.append(rec)
        name = f"{rec.get('First Name','')} {rec.get('Last Name','')}".strip()
        print(name or "(name not found)")
        time.sleep(1.0)

    driver.quit()
    print("\n📊  Writing Excel …")
    save_excel(records, OUTPUT_FILE)
    print("🎉  Done!")